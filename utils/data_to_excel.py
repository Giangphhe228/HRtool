import pandas as pd
import os
import openpyxl
import xlsxwriter
from copy import copy
from django.db import connection, connections
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import range_boundaries
from urllib.parse import urlparse
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import Text

levels = {
                2: 'L1',
                3: 'L2',
                4: 'L3',
                # -1: 'NoLevel',
                1: 'SVCNTS'
                }
# tạo biến tên cho các trường trong dataframe để sử dụng sau
krid ='krId'
krp = 'KR phòng'
krt = 'KR team'
krc = 'KR cá nhân'
ctt = 'Công thức tính'
ei = 'employeeId'
email = 'email'
# trường đang dùng để mapping các kpi với nhau
type = 'Loại'
name = 'Name'
lv = 'level'
tn = 'teamName'
note = 'Note dự kiến'
noteproof = 'Note bằng chứng thực tế'
QA = 'QA'
proofStr = 'Bằng chứng'
proofURL ='Link'
tsct = '% Trọng số chỉ tiêu'
kq = 'Kết quả'
tl = 'Tỷ lệ'
et = 'Tổng thời gian dự kiến/ ước tính công việc (giờ)'
rt = 'Tổng thời gian thực hiện công việc thực tế (giờ)'

def is_integer(input_str):
    try:
        return int(input_str.replace('"',''))  
    except ValueError:
        raise ValueError(f"'{input_str}' cannot be converted to an integer.")  # Nếu xảy ra lỗi, raise ValueError

def is_valid_string(s):
    if isinstance(s, str):
        return True
    else:
        return False

def extract_month_year(date_string):
    parts = date_string.split('/')
    if len(parts) == 2:
        month = int(parts[0].replace('"',''))
        year = int(parts[1].replace('"',''))
        return month, year
    else:
        raise ValueError("Invalid date format. Expected MM/YYYY.")

def GenerateExcelSheet(basedir,data_dictionary) -> None:
        # Đóng kết nối hiện tại
        connection.close()
        # Mở kết nối mới với cơ sở dữ liệu khác (thay 'replica' bằng alias mong muốn)
        new_db_connection = connections['fedatabase']
        cursor = new_db_connection.cursor()
        sql_query='''select 	        emp.employee_code as employeeId,
                                        emp.full_name as name,
                                        emp.email_address as email,
                                        lv.name as level,
                                        tm.id as teamId,
                                        tm.name as teamName,
                                        okr.id as krId,
                                        okr.type,
                                        okr.objective_name as krDep,
                                        okr.title as krTeam,
                                        okr.key_result_name as krPer,
                                        okr.formula as formulaName,
                                        okr.source as sourceName,
                                        okr.regularly,
                                        okr.unit,
                                        okr.condition,
                                        okr.norm,
                                        okr.weight,
                                        okr.result,
                                        okr.ratio,
                                        okr.estimate_time,
                                        okr.actual_time,
                                        okr.estimate_time_note,
                                        okr.note,
                                        okr.qa_Status,
                                        pr.name,
		                                pr.url 
                                from "employees" as emp
                                LEFT JOIN "employees_team_links" as etl ON emp.id=etl.employee_id 
                                LEFT JOIN "teams" as tm ON etl.team_id=tm.id
                                INNER JOIN "okr_kpis_employee_links" as okel ON emp.id=okel.employee_id 
                                INNER JOIN "okr_kpis" as okr ON okel.okr_kpi_id=okr.id
                                LEFT JOIN "employees_department_links" as edl ON emp.id=edl.employee_id
                                LEFT JOIN "departments" as de ON edl.department_id=de.id
                                LEFT JOIN "employees_level_links" as ell ON emp.id = ell.employee_id
                                LEFT JOIN "levels" as lv ON ell.level_id = lv.id
                                LEFT JOIN "proofs_okr_kpi_links" as pokl ON okr.id=pokl.okr_kpi_id
                                LEFT JOIN "proofs" as pr ON pokl.proof_id=pr.id
                                WHERE 1=1 and okr.regularly = 'month'
                    '''
        # print("month", type(data_dictionary.get("month")))
        # print("year", type(data_dictionary.get("year")))
        # print("department_id", type(data_dictionary.get("department_id")))
        # Kiểm tra và thêm điều kiện từ dictionary
        if data_dictionary.get("month") is not None:
            sql_query += " and EXTRACT(MONTH FROM okr.created_at) = %(month)s"
        if data_dictionary.get("year") is not None:
            sql_query += " and EXTRACT(YEAR FROM okr.created_at) = %(year)s"
        if data_dictionary.get("department_id") is not None:
            sql_query += " and de.id = %(department_id)s"
        # print("đay là data_dictionary: ", data_dictionary)
        cursor.execute(sql_query, data_dictionary)
        result = cursor.fetchall()
        dataframe = pd.DataFrame(result)
        # đóng kết nối sau khi đã lấy được dữ liệu
        new_db_connection.close()
        cursor.close()
        dataframe.columns = ['employeeId', 'Name', 'email', 'level', 'teamId', 'teamName',
                             'krId','Loại', 'KR phòng', 'KR team', 'KR cá nhân', 'Công thức tính',
                            'Nguồn dữ liệu', 'Định kỳ tính', 'Đơn vị tính', 'Điều kiện', 'Norm',
                            '% Trọng số chỉ tiêu', 'Kết quả', 'Tỷ lệ', 'Tổng thời gian dự kiến/ ước tính công việc (giờ)',
                                'Tổng thời gian thực hiện công việc thực tế (giờ)', 'Note dự kiến','Note bằng chứng thực tế','QA','Bằng chứng','Link']

        dataframe.sort_values( ei, inplace=True)
        # dataframe.drop(columns=['krId'], axis=1, inplace=True)
        dataframe.drop(columns=['teamId'], axis=1, inplace=True)
        # dataframe.replace("NUM", "%", inplace=True)
        # dataframe.replace("CAT", "Đạt/Không đạt", inplace=True)
        dataframe.replace("month", "Tháng", inplace=True)
        # dataframe.fillna(0, inplace=True)
        formatResultExcelSheet(dataframe)
        for value, str in levels.items():
            temp_df = dataframe.loc[dataframe['level'] == str]
            if temp_df.empty:
                    raise ValueError("một trường quan trọng đang ko có đầu vào dẫn đến trả về dataframe trống không thể convert sang excel(level,department_id,deadline...)")
            # temp_df.drop(columns=['level'], axis=1)
            file_name=f"/nhóm {str}.xlsx"
            file_directory = basedir+file_name
            if os.path.exists(file_directory):
                    os.remove(file_directory)
            # print("đây là temp_df:\n",temp_df)
            temp_df.to_excel(file_directory)
            # formatKPIExcelSheetWithXLSXWriter(file_directory,str)
            formatKPIExcelSheet(file_directory,str)

def excelToDataframe(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    data = sheet.values
    columns = next(data)  # Lấy tên cột từ hàng đầu tiên
    # Tạo DataFrame từ dữ liệu
    df = pd.DataFrame(data, columns=columns)
    wb.close()
    return df

def assignHyperlinksForCellByCoordination(sheet, name_pos, link_pos, note_pos, file_path):
    # Đọc nội dung của sheet thành DataFrame
    data = sheet.values
    columns = next(data)  # Lấy tên cột từ dòng đầu tiên
    convert_df = pd.DataFrame(data, columns=columns)  
    # print("đây là convert_df : \n",convert_df)  

    xlsxworkbook = xlsxwriter.Workbook(options={'nan_inf_to_errors': True})
    xlsxsheet = xlsxworkbook.add_worksheet()

    for row_num, row_data in enumerate(convert_df):
        xlsxsheet.write_row(row_num, 0, row_data)

    hyperlink_format = xlsxworkbook.add_format({
    'color': 'blue',
    'underline': 1,
    })
    
    row_index = 0
    for row in convert_df:
        text_string = row[name_pos]  
        url_string = row[link_pos]   
        row_index += 1 
        if is_valid_string(url_string):      
            if url_string not in proofURL:
                cell_values = text_string.split(',')  # Split the comma-separated values
                urls = url_string.split(',')
                # print("đây là urls : \n",urls)
                for value, url in zip(cell_values,urls):
                    rich_text = xlsxworkbook._xml_rich_inline_string
                    rich_text.append(value)
                    rich_text.append()
                    # cell_value = f'=HYPERLINK({url}, {value})'
                xlsxsheet.write(row_index + 3, note_pos+1, rich_text) 

    xlsxworkbook.close(file_path)

# format ra file kpi có các sheet kpi input là path các file excel
def formatKPIExcelSheet(file_path,level) -> None:
# xử lý dữ liệu với dataframe của pandas 
# -----------------------------------------------------------------------------------------------------------
    # Tạo DataFrame từ dữ liệu
    df = excelToDataframe(file_path)
    print("đây là file_path: \n",file_path)
    print("đây là df[tsct]: \n",df[tsct])
    df[tsct].fillna(0, inplace=True)
    df[kq].fillna(0, inplace=True)
    df[tl].fillna(0, inplace=True)
    df[et].fillna(0, inplace=True)
    df[rt].fillna(0, inplace=True)

    df[tsct] = df[tsct].astype(int)
    df[kq] = df[kq].astype(float)
    df[tl] = df[tl].astype(float)
    df[tl]=(df[tsct]*df[kq])/100
    df[et] = df[et].astype(float)
    df[rt] = df[rt].astype(float)
    df[email] = df[email].apply(lambda x: x.split('@')[0] if x is not None else x)
    df[QA] = df[QA].apply(lambda x: 'OK' if x is True else ('NOT OK' if x is False else ''))
    df[ei] = df[ei].apply(lambda x: '' if x == None else x)
    df[proofURL].fillna('', inplace=True)
    df[proofStr].fillna('', inplace=True)
    df[proofStr] = df[proofStr].apply(lambda x: '' if x == None else x)
    df[proofURL] = df[proofURL].apply(lambda x: '' if x == None else x)
    df[type] = df[type].apply(lambda x: '' if x == None else x)
    # tính tổng các trường cần tính
    agg_funcs = {proofStr: ','.join, proofURL: ','.join}
    proof_sum = df.groupby( krid, as_index=False).agg(agg_funcs)
    proof_sum[proofURL].fillna('', inplace=True)
    proof_sum[proofStr].fillna('', inplace=True)
    proof_sum[proofStr] = proof_sum[proofStr].apply(lambda x: '' if x == None else x)
    proof_sum[proofURL] = proof_sum[proofURL].apply(lambda x: '' if x == None else x)
    df.drop(proofStr, axis=1, inplace=True)
    df.drop(proofURL, axis=1, inplace=True)
    df  = pd.merge(df.drop_duplicates(subset=krid), proof_sum, on=krid)
    tsct_sum = df.groupby([ei, type])[tsct].sum().reset_index()
    kq_sum = df.groupby([ei, type])[kq].sum().reset_index()
    tl_sum = df.groupby([ei, type])[tl].sum().reset_index()
    et_sum = df.groupby([ei])[et].sum().reset_index()
    rt_sum = df.groupby([ei])[rt].sum().reset_index()
    # Tạo DataFrame chứa các tổng
    df_data_sum = pd.DataFrame({ei: tsct_sum[ei],
                           type: tsct_sum[type],
                           tsct: tsct_sum[tsct],
                           kq: kq_sum[kq],
                           tl: tl_sum[tl]})
    df_time_sum = pd.DataFrame({ei: et_sum[ei],
                           et: et_sum[et],
                           rt: rt_sum[rt]})

    # merge các tổng đã tính ở trên vào một hàng và tìm vị trị cần điền các tổng đó trong dataframe gốc
    index_df_data_sum = df_data_sum[[ei, type]]
    # Tạo một cột boolean cho biết các hàng có là bản sao của hàng trước đó hay không
    is_duplicated = index_df_data_sum.duplicated(subset=[ei, type], keep='last')
    # Chỉ giữ lại các hàng cuối cùng của mỗi cặp giá trị bằng nhau
    max_indices = index_df_data_sum[~is_duplicated]
    # print("đây là max_indices: \n",max_indices)
    merged_sum_idx_df = pd.merge(
        max_indices, df_time_sum, on=ei, how='left')
    # print("đây là merged_sum_idx_df",merged_sum_idx_df)
    merged_sum_idx_df.drop(columns=[ei], axis=1, inplace=True)
    merged_sum_idx_df[et] = merged_sum_idx_df[et].where(merged_sum_idx_df[type] != "kpi")
    merged_sum_idx_df[rt] = merged_sum_idx_df[rt].where(merged_sum_idx_df[type] != "kpi")
    # merged_sum_idx_df[et] = merged_sum_idx_df[type].apply(lambda merged_sum_idx_df: '' if merged_sum_idx_df[type] == "kpi" else merged_sum_idx_df[et])
    # merged_sum_idx_df[rt] = merged_sum_idx_df[type].apply(lambda merged_sum_idx_df: '' if merged_sum_idx_df[type] == "kpi" else merged_sum_idx_df[rt])
    merged_sum_df = pd.merge(df_data_sum, merged_sum_idx_df,
                             left_index=True, right_index=True, how='left')
    # print("đây là merged_sum_df",merged_sum_df)
    merged_sum_df.fillna(0, inplace=True)

    merged_sum_df[et] = merged_sum_df[et].apply(lambda x: '' if x == 0 else x)
    merged_sum_df[rt] = merged_sum_df[rt].apply(lambda x: '' if x == 0 else x)
    # sắp xếp lại và lấy vị trí các đoạn cần insert các tổng vào (need_add_index_df), đồng thời dùng sorted_df cho các thao tác sau (sorted_df)
    sorted_df = df.sort_values(by=[ei, type], ascending=[
                               True, True]).reset_index()
    # xóa các record duplicate nếu 1 kpi/okr có nhiều proof
    sorted_df.drop(columns=['index'], axis=1, inplace=True)
    sorted_df.drop(columns=[krid], axis=1, inplace=True)
    # print("dataframe sorted_df: \n",sorted_df)
    index_sorted_df = sorted_df[[ei, type]]
    # Tạo một cột boolean cho biết các hàng có là bản sao của hàng trước đó hay không
    is_duplicated = index_sorted_df.duplicated(subset=[ei, type], keep='last')
    # Chỉ giữ lại các hàng cuối cùng của mỗi cặp giá trị bằng nhau
    need_add_index_df = index_sorted_df[~is_duplicated]
    # print("đây là need_add_index_df : \n",need_add_index_df)
    # thêm các kết quả đã tính toán ở trên vào cuối của mỗi kpi
    added_row = 0
    map_new_index = {}
    for (index, i) in zip(need_add_index_df.index, range(len(merged_sum_df))):
            df_new_record = pd.DataFrame(merged_sum_df.iloc[i, :]).T
            true_position = index+added_row+1
            sorted_df = pd.concat([sorted_df.iloc[:true_position], df_new_record,
                                  sorted_df.iloc[true_position:]]).reset_index(drop=True)
            # thêm các giá trị index mới vào dictionary
            map_new_index[index] = true_position
            added_row += 1
    sorted_df.drop(columns=[None], axis=1, inplace=True)
    sorted_df = sorted_df.apply(lambda x: '' if x.empty else x)
    sorted_df.drop(columns=[type+"_x"], axis=1, inplace=True)
    sorted_df.drop(columns=[type+"_y"], axis=1, inplace=True)
    # thay đổi thành rỗng để có thể groupby phía sau
    sorted_df[tn] = sorted_df[tn].apply(lambda x: '' if x == None else x)
    need_add_index_df.set_index(need_add_index_df.index.map(map_new_index), inplace=True)

    # 1. tạo ra các dataframe chứa tên, level, tên nhóm và dataframe chưa tên các cột  (user_df)
    grouped = sorted_df.groupby([ei, name,email, lv, tn])
    user_df = grouped.size().reset_index(name='Count')
    user_df.insert(0, 'Index', user_df.index+1)
    user_df.drop(columns=['Count'], axis=1, inplace=True)
    # bỏ cột id trong giao diện
    user_df.drop(columns=[ei], axis=1, inplace=True)
    # print("đây là user_df : \n",user_df)

    # Create a new DataFrame to store the column names (title_df)'
    columns_to_drop = [ei, name,email, lv, tn]
    insert_header_name_df = sorted_df.drop(columns=columns_to_drop, axis=1)
    title_df = pd.DataFrame([insert_header_name_df.columns],
                            columns=insert_header_name_df.columns)
    # print("đây là user_df : \n",title_df)

    # 2. tạo ra một list chứa các vị trí cần add các record đó vào bằng cách tìm các vị trí đầu xuất hiện của các record
    # Tạo cột mới để xác định lần xuất hiện đầu tiên của giá trị trùng nhau
    sorted_df['FirstOccurrence'] = ~sorted_df.duplicated(subset=[ei])
    # lấy lần xuất hiện đầu tiên của các giá trị trùng nhau (firstOccurrence_df)
    firstOccurrence_df = sorted_df[sorted_df['FirstOccurrence']].drop(
        columns='FirstOccurrence')
    firstOccurrence_list = firstOccurrence_df.index
    sorted_df = sorted_df.drop(columns='FirstOccurrence', axis=1)


    # Chuyển DataFrame thành một đối tượng Sheet bằng pandas dataframe_to_rows
    user_rows = list(dataframe_to_rows(user_df, index=False, header=False))
    if len(user_rows) != len(firstOccurrence_list):
        raise ValueError("Số lượng dòng cần thêm phải bằng số dự liệu muốn thêm!")
    
    # tạo ra 1 dataframe có nhiều hàng là tên các cột cần add vào sheet
    for i in range(len(user_rows)):
         title_df = pd.concat([title_df, title_df], ignore_index=True)

    # lấy vị trí các cột cần mở rộng khi align bằng openpyxl
    sum_column_names = [tsct, kq, tl, et,rt]
    sum_column_positions = [title_df.columns.get_loc(col_name) for col_name in sum_column_names]

    column_name_rows = list(dataframe_to_rows(title_df, index=False, header=False))
    rows = dataframe_to_rows(insert_header_name_df, index=False, header=False)


# format các trường dữ liệu với workbook của openpyxl
# -----------------------------------------------------------------------------------------------------------
    # tạo một sheet excel mới
    workbook = Workbook()
    sheet = workbook.active
    # Ghi dữ liệu từ DataFrame vào Workbook
    # Gán tiêu đề cho sheet
    Header_font = Font(name='Arial',size=16, bold=True)
    header_text=f"nhóm {level}"
    sheet.cell(row=1, column=1, value=header_text).font = Header_font
    # Merge các ô trong dòng đầu tiên
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_df.shape[1])

    # Chèn các dòng vào vị trí xác định (từ dòng 2 trở đi)
    data_font = Font(name='Arial',size=11, bold=False)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)
            sheet.cell(row=r_idx, column=c_idx, value=value).font = data_font

    # thêm các dòng thông tin người dung và title của các trường vào sheet, đồng thời thêm màu, sửa font
    # light_blue_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
    light_blue_fill = PatternFill(start_color='CCE0FF', end_color='CCE0FF', fill_type='solid')
    alittler_dark_blue_fill = PatternFill(start_color='9FC9E7', end_color='9FC9E7', fill_type='solid')
    dark_blue_fill = PatternFill(start_color='6699CC', end_color='6699CC', fill_type='solid')
    light_yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    light_green_fill = PatternFill(start_color='B0E57C', end_color='B0E57C', fill_type='solid')
    dark_green_fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
    user_title_font = Font(name='Arial',size=11, bold=True,color="FF0000")
    title_font = Font(name='Arial',size=11, bold=True,color="FFFFFF")
    
    # lấy vị trí các column cần đổi màu xanh lá cây nhạt
    column_light_green_fill = [tsct, kq, tl, et, note]
    column_positions_light_green_fill = [title_df.columns.get_loc(col_name)+1 for col_name in column_light_green_fill]

    # lấy vị trí các column cần đổi màu xanh lá cây đậm
    column_dark_green_fill = [QA]
    column_positions_dark_green_fill = [title_df.columns.get_loc(col_name)+1 for col_name in column_dark_green_fill]

    # # lấy vị trí các column cần đổi màu xanh lá cây đậm
    # column_dark_green_fill = [tsct, 'KR team', 'KR cá nhân', 'Công thức tính', note]
    # column_positions_dark_green_fill = [title_df.columns.get_loc(col_name) for col_name in column_dark_green_fill]

    # lấy vị trí các column cần đổi màu xanh da trời nhạt
    column_light_blue_fill = [type,rt,noteproof]
    column_positions_light_blue_fill = [title_df.columns.get_loc(col_name)+1 for col_name in column_light_blue_fill]

    for row_sum_index in need_add_index_df.index:
        for col_sum_index in sum_column_positions:
            # print("col and row",col_sum_index,row_sum_index)
            # ở đây row phải +2 là vì: sheet có các index từ số 1 + header là lấy mất 1 dòng đầu
            # ở đây col phải +1 là vì: sheet có các index từ số 1
            sheet.cell(row=row_sum_index+2, column=col_sum_index+1).fill = light_yellow_fill

    added_sheet_row = 0
    for insert_index, row_user_data,row_column_name in zip(firstOccurrence_list, user_rows,column_name_rows):       
        user_insert = insert_index+added_sheet_row+2
        title_insert = insert_index+added_sheet_row+3
        sheet.insert_rows(user_insert, 1)
        sheet.insert_rows(title_insert, 1)
        # thêm dòng thông tin user
        for col_idx_user,user_value in enumerate(row_user_data, 1):
            sheet.cell(row=user_insert, column=col_idx_user, value=user_value)
            sheet.cell(row=user_insert, column=col_idx_user).fill = light_blue_fill
            sheet.cell(row=user_insert, column=col_idx_user).font = user_title_font
        # thêm dòng title
        for col_idx,column_name_value in enumerate(row_column_name,1):
            sheet.cell(row=title_insert, column=col_idx, value=column_name_value)
            sheet.cell(row=title_insert, column=col_idx).fill = dark_blue_fill
            sheet.cell(row=title_insert, column=col_idx).font = title_font
            
            for green_index in column_positions_light_green_fill:
                if green_index==col_idx:
                    # print("vị trí xanh lá cây:",green_index)
                    sheet.cell(row=title_insert, column=green_index).fill = light_green_fill

            for dark_green_index in column_positions_dark_green_fill:
                if dark_green_index==col_idx:
                    # print("vị trí xanh lá cây đậm:",dark_green_index)
                    sheet.cell(row=title_insert, column=dark_green_index).fill = dark_green_fill

            for blue_index in column_positions_light_blue_fill:
                if blue_index==col_idx:
                    # print("vị trí xanh da trời nhạt:",blue_index)
                    sheet.cell(row=title_insert, column=blue_index).fill = alittler_dark_blue_fill

        if(insert_index!=0):
             blank_insert = insert_index+added_sheet_row+2
             sheet.insert_rows(blank_insert, 1) 
             added_sheet_row+=3       
        else:
             added_sheet_row+=2

    # Biến trạng thái
    merging = False
    merge_start = None
    # lấy vị trí cột loại để merge các cell 
    merge_column_position = title_df.columns.get_loc(type)

    # Duyệt qua các dòng trong cột "loại"
    for row_index, row in enumerate(sheet.iter_rows(min_col=merge_column_position, max_col=merge_column_position, values_only=True), start=1):
        if row[0] =='kpi' or row[0]=='okr':
            if not merging:
                merging = True
                merge_start = row_index
        else : 
            if merging:
                merging = False
                sheet.merge_cells(start_row=merge_start, start_column=1, end_row=row_index - 1, end_column=1)


    # lấy vị trí column link bằng chứng
    note_proof_column = [noteproof,proofStr,proofURL]
    note_proof_column_position = [title_df.columns.get_loc(col_name) for col_name in note_proof_column]

    row_index = 0
    font = Font(name='Arial',
             size=11,
             bold=False,
             italic=False,
             vertAlign=None,
             underline='none',
             strike=False,
             color='00FF0000')
    inline_font = InlineFont(font)

    url_font = Font(name='Arial',
             size=11,
             bold=False,
             italic=False,
             vertAlign=None,
             underline='none',
             strike=False,
             color='6699CC')
    url_inline_font = InlineFont(url_font)

    for row in sheet.iter_rows(min_row=4, values_only=True):
        text_string = row[note_proof_column_position[1]]  
        url_string = row[note_proof_column_position[2]]   
        row_index += 1 
        if is_valid_string(url_string):      
            if url_string not in proofURL:
                rich_text = CellRichText()
                cell_values = text_string.split(',')  # Split the comma-separated values
                urls = url_string.split(',')
                # print("đây là urls : \n",urls)
                for value, url in zip(cell_values,urls):
                    text_block = TextBlock(text=value+" : ",font=inline_font)
                    url_block = TextBlock(text=url+"\n",font=url_inline_font)
                    rich_text.append(text_block)
                    rich_text.append(url_block)
                    # cell_value = f'=HYPERLINK({url}, {value})'
                cell = sheet.cell(row=row_index + 3, column=note_proof_column_position[0]+1) 
                cell.value = rich_text
                # print("đây là cell : \n",cell)
    # Xóa cột dựa trên vị trí
    sheet.delete_cols(note_proof_column_position[1]+1, note_proof_column_position[2]+1)

    # Thiết lập tự động tăng kích thước các cột
    # for col in sheet.columns:
    #     max_length = 0
    #     for cell in col:
    #         try:
    #             if len(str(cell.value)) > max_length:
    #                 max_length = len(cell.value)
    #         except:
    #             pass
    #     adjusted_width = (max_length + 2) * 1.2
    #     sheet.column_dimensions[col[0].column_letter].width = adjusted_width

    # Đặt viền cho các cell
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    # Loop through each worksheet in the workbook
    for sheet in workbook.worksheets:
        # Loop through each row in the worksheet
        for row in sheet.iter_rows():
            # Loop through each cell in the row
            for cell in row:
                # Set wrap_text to True to automatically wrap text
                cell.alignment = openpyxl.styles.Alignment(wrapText=True)
                # thêm viền
                cell.border = border
                # Thiết lập tự động align vào giữa cho toàn bộ text trong sheet
                cell.alignment = Alignment(wrap_text=True,horizontal='center', vertical='center')

        # Auto-adjust row height for all rows in the worksheet
        for row in sheet.iter_rows():
            sheet.row_dimensions[row[0].row].auto_size = True
        
    # Define a dictionary to store the desired column widths
    # column_widths = {'KR phòng': 20, 'KR team': 20, 'KR cá nhân': 20, 'Note': 20}  

    # lấy vị trí các cột cần mở rộng khi align bằng openpyxl
    column_names_30 = [krp, krt, krc, ctt, note,noteproof,QA]
    column_positions_30 = [title_df.columns.get_loc(col_name) for col_name in column_names_30]

    column_names_20 = [ type, et, rt]
    column_positions_20 = [title_df.columns.get_loc(col_name) for col_name in column_names_20]

    # Update the column widths
    for index in column_positions_30:
        # đoạn này index +1 là vì vào sheet excel các index tính từ 1 chứ không còn là 0 như ở dataframe
        column_letter = openpyxl.utils.get_column_letter(index+1)
        sheet.column_dimensions[column_letter].width = 35
    # Update the column widths
    for index in column_positions_20:
        # đoạn này index +1 là vì vào sheet excel các index tính từ 1 chứ không còn là 0 như ở dataframe
        column_letter = openpyxl.utils.get_column_letter(index+1)
        sheet.column_dimensions[column_letter].width = 20
        
    # Đọc nội dung của sheet thành DataFrame
    # data = sheet.values
    # columns = next(data)  # Lấy tên cột từ dòng đầu tiên
    # final_df = pd.DataFrame(data, columns=columns)  
    # print("đây là excel file cuối : \n",final_df)

    # Lưu Workbook
    workbook.close()
    workbook.save(file_path)

def synthesizeExcelFilebySheet(listDirectory,targetDirectory) -> None:

    # Tạo một workbook mới để tổng hợp dữ liệu
    wb_combined = openpyxl.Workbook()
    # Lặp qua từng file Excel gốc
    for file_name in listDirectory:
        # Mở file Excel gốc
        wb_original = openpyxl.load_workbook(file_name)
        # Chọn sheet trong file Excel gốc 
        sheet_original = wb_original.active 
        for value,levelstr in levels.items():
            # Tạo một sheet mới trong workbook tổng hợp có title phù hợp với level của nhân viên trong sheet
            title=f"nhóm {levelstr}"
            if title in file_name:   
                sheet_combined = wb_combined.create_sheet(title)
            else:
                continue        
        # Lặp qua từng hàng và cột trong sheet gốc
        for row in sheet_original.iter_rows():
            for cell in row:
                # Copy giá trị sang sheet tổng hợp
                new_cell = sheet_combined.cell(row=cell.row, column=cell.column, value=cell.value)
                
                # Sao chép các thuộc tính định dạng
                new_cell.font = openpyxl.styles.Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=cell.font.color,
                    underline=cell.font.underline
                )
                
                new_cell.fill = openpyxl.styles.PatternFill(
                    fill_type=cell.fill.fill_type,
                    start_color=cell.fill.start_color,
                    end_color=cell.fill.end_color
                )
                
                new_cell.border = openpyxl.styles.Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                
                new_cell.alignment = openpyxl.styles.Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    text_rotation=cell.alignment.text_rotation,
                    wrap_text=cell.alignment.wrap_text,
                    shrink_to_fit=cell.alignment.shrink_to_fit,
                    indent=cell.alignment.indent
                )

                new_cell.hyperlink = cell.hyperlink

        # Sao chép độ rộng cột từ sheet gốc sang sheet tổng hợp
        for col in sheet_original.column_dimensions:
            sheet_combined.column_dimensions[col] = sheet_original.column_dimensions[col]

        # Lấy danh sách các vùng merge
        merged_ranges = sheet_original.merged_cells.ranges
        # Loop through each merged cell range
        for merged_range in merged_ranges:
            # Extract the start and end row indices from the range string
            start_col, start_row, end_col, end_row = range_boundaries(merged_range.coord)      
            # Merge the cells in the range
            sheet_combined.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    # Xóa sheet mặc định trong Workbook tổng hợp
    if 'Sheet' in wb_combined.sheetnames:
        wb_combined.remove(wb_combined['Sheet'])
    # Lưu workbook tổng hợp lại vào một file mới
    wb_combined.save(targetDirectory)

def formatResultExcelSheet(dataframe) -> None:
    result_sheet_df=dataframe[[ei, name, lv, tn, type, kq, tl, et, rt]]
    result_sheet_df.sort_values( ei, inplace=True)
    result_sheet_df[kq] = result_sheet_df[kq].astype(float)
    result_sheet_df[tl] = result_sheet_df[tl].astype(float)
    result_sheet_df[et] = result_sheet_df[et].astype(float)
    result_sheet_df[rt] = result_sheet_df[rt].astype(float)
    kq_sum = result_sheet_df.groupby([ei, type])[kq].sum().reset_index()
    tl_sum = result_sheet_df.groupby([ei, type])[tl].sum().reset_index()
    et_sum = result_sheet_df.groupby([ei])[et].sum().reset_index()
    rt_sum = result_sheet_df.groupby([ei])[rt].sum().reset_index()
    result_sheet_df.drop(columns=[type], axis=1, inplace=True)
    result_sheet_df.drop(columns=[kq], axis=1, inplace=True)
    result_sheet_df.drop(columns=[tl], axis=1, inplace=True)
    result_sheet_df.drop(columns=[et], axis=1, inplace=True)
    result_sheet_df.drop(columns=[rt], axis=1, inplace=True)
    df_data_sum = pd.DataFrame({ei: tl_sum[ei],
                                type: tl_sum[type],
                                kq: kq_sum[kq],
                           tl: tl_sum[tl]})
    df_time_sum = pd.DataFrame({ei: et_sum[ei],
                           et: et_sum[et],
                           rt: rt_sum[rt]})
    
    list_main_part = result_sheet_df.groupby([ei, name,lv]).apply(list)
    main_part_df= list_main_part.reset_index(name="index")
    main_part_df.drop(columns=['index'], axis=1, inplace=True)
    df_data_sum_kpi=df_data_sum[df_data_sum[type] == 'kpi']
    df_data_sum_kpi.drop(columns=[kq], axis=1, inplace=True)
    df_data_sum_kpi.drop(columns=[type], axis=1, inplace=True)
    df_data_sum_kpi = df_data_sum_kpi.rename(columns={tl: 'kết quả KPI'}) 
    # print("đây là df_data_sum_kpi:\n",df_data_sum_kpi)

    df_data_sum_okr=df_data_sum[df_data_sum[type] == 'okr'] 
    df_data_sum_okr.drop(columns=[kq], axis=1, inplace=True)  
    df_data_sum_okr.drop(columns=[type], axis=1, inplace=True)
    df_data_sum_okr = df_data_sum_okr.rename(columns={tl: 'kết quả OKR'})
    # print("đây là df_data_sum_okr:\n",df_data_sum_okr)

    merged_df =pd.merge(main_part_df,df_data_sum_kpi, on=ei, how='left')
    kpi_merged_df=pd.merge(merged_df,df_data_sum_okr, on=ei, how='left')
    kpi_merged_df['NSNS']=''
    kpi_merged_df['Số giờ OT']=''
    time_merged_df=pd.merge(kpi_merged_df,df_time_sum, on=ei, how='left')
    time_merged_df['Thời gian chênh lệch']= time_merged_df[rt]-time_merged_df[et]
    time_merged_df.rename_axis('STT')
    time_merged_df.rename(columns={ei: "Mã NV"}, inplace=True)
    time_merged_df.rename(columns={name: "Họ tên"}, inplace=True)
    time_merged_df.rename(columns={lv: "Level"}, inplace=True)
    

    for value,levelstr in levels.items():
        # globals()[levelstr+"_df"]
        df = time_merged_df[time_merged_df["Level"] == levelstr].reset_index()
        df.drop(columns=["Mã NV"], axis=1, inplace=True)
        df.drop(columns=["Level"], axis=1, inplace=True)
        df.drop(columns=["index"], axis=1, inplace=True)
        df.rename(columns={"Họ tên": "Nhân sự"}, inplace=True)
        kpi_mean = df['kết quả KPI'].mean()
        okr_mean = df['kết quả OKR'].mean()
        et_mean = df[et].mean()
        rt_mean = df[rt].mean()
        cl_mean = df['Thời gian chênh lệch'].mean()
        print("kpi_mean:\n",kpi_mean)
        df_data_mean = pd.DataFrame({ "Nhân sự": ["Trung bình"],
                                'kết quả KPI': [kpi_mean],
                                'kết quả OKR': [okr_mean],
                                et: [et_mean],
                                rt: [rt_mean],
                                'Thời gian chênh lệch': [cl_mean]})
        globals()[levelstr+"_df"] = pd.concat([df, df_data_mean], ignore_index=True)
        print("df:")
        print(df)
        
def formatKPIExcelSheetWithXLSXWriter(file_path,level) -> None:
# xử lý dữ liệu với dataframe của pandas 
# -----------------------------------------------------------------------------------------------------------
    # Tạo DataFrame từ dữ liệu
    df = excelToDataframe(file_path)
    df[tsct] = df[tsct].astype(int)
    df[kq] = df[kq].astype(float)
    df[tl] = df[tl].astype(float)
    df[et] = df[et].astype(float)
    df[rt] = df[rt].astype(float)
    df[email] = df[email].apply(lambda x: x.split('@')[0] if x is not None else x)
    df[QA] = df[QA].apply(lambda x: 'OK' if x is True else ('NOT OK' if x is False else 'Trống'))
    df[ei] = df[ei].apply(lambda x: '' if x == None else x)
    df[proofURL].fillna('', inplace=True)
    df[proofStr].fillna('', inplace=True)
    df[proofStr] = df[proofStr].apply(lambda x: '' if x == None else x)
    df[proofURL] = df[proofURL].apply(lambda x: '' if x == None else x)
    df[type] = df[type].apply(lambda x: '' if x == None else x)
    # tính tổng các trường cần tính
    agg_funcs = {proofStr: ','.join, proofURL: ','.join}
    proof_sum = df.groupby( krid, as_index=False).agg(agg_funcs)
    proof_sum[proofURL].fillna('', inplace=True)
    proof_sum[proofStr].fillna('', inplace=True)
    proof_sum[proofStr] = proof_sum[proofStr].apply(lambda x: '' if x == None else x)
    proof_sum[proofURL] = proof_sum[proofURL].apply(lambda x: '' if x == None else x)
    tsct_sum = df.groupby([ei, type])[tsct].sum().reset_index()
    kq_sum = df.groupby([ei, type])[kq].sum().reset_index()
    tl_sum = df.groupby([ei, type])[tl].sum().reset_index()
    et_sum = df.groupby([ei])[et].sum().reset_index()
    rt_sum = df.groupby([ei])[rt].sum().reset_index()
    # Tạo DataFrame chứa các tổng
    df_data_sum = pd.DataFrame({ei: tsct_sum[ei],
                           type: tsct_sum[type],
                           tsct: tsct_sum[tsct],
                           kq: kq_sum[kq],
                           tl: tl_sum[tl]})
    df_time_sum = pd.DataFrame({ei: et_sum[ei],
                           et: et_sum[et],
                           rt: rt_sum[rt]})

    # merge các tổng đã tính ở trên vào một hàng và tìm vị trị cần điền các tổng đó trong dataframe gốc
    index_df_data_sum = df_data_sum[[ei, type]]
    # Tạo một cột boolean cho biết các hàng có là bản sao của hàng trước đó hay không
    is_duplicated = index_df_data_sum.duplicated(subset=[ei, type], keep='last')
    # Chỉ giữ lại các hàng cuối cùng của mỗi cặp giá trị bằng nhau
    max_indices = index_df_data_sum[~is_duplicated]
    # print("đây là max_indices: \n",max_indices)
    merged_sum_idx_df = pd.merge(
        max_indices, df_time_sum, on=ei, how='left')
    merged_sum_idx_df.drop(columns=[ei], axis=1, inplace=True)
    merged_sum_df = pd.merge(df_data_sum, merged_sum_idx_df,
                             left_index=True, right_index=True, how='left')
    merged_sum_df.fillna(0, inplace=True)

    merged_sum_df[et] = merged_sum_df[et].apply(lambda x: '' if x == 0 else x)
    merged_sum_df[rt] = merged_sum_df[rt].apply(lambda x: '' if x == 0 else x)
    # print("đây là merged_sum_df: \n",merged_sum_df)



    # sắp xếp lại và lấy vị trí các đoạn cần insert các tổng vào (need_add_index_df), đồng thời dùng sorted_df cho các thao tác sau (sorted_df)
    sorted_df = df.sort_values(by=[ei, type], ascending=[
                               True, True]).reset_index()
    # xóa các record duplicate nếu 1 kpi/okr có nhiều proof
    sorted_df.drop(columns=['index'], axis=1, inplace=True)
    sorted_df.drop(proofStr, axis=1, inplace=True)
    sorted_df.drop(proofURL, axis=1, inplace=True)
    sorted_df  = pd.merge(sorted_df.drop_duplicates(subset=krid), proof_sum, on=krid)
    sorted_df.drop(columns=[krid], axis=1, inplace=True)
    # print("dataframe sorted_df: \n",sorted_df)
    index_sorted_df = sorted_df[[ei, type]]
    # Tạo một cột boolean cho biết các hàng có là bản sao của hàng trước đó hay không
    is_duplicated = index_sorted_df.duplicated(subset=[ei, type], keep='last')
    # Chỉ giữ lại các hàng cuối cùng của mỗi cặp giá trị bằng nhau
    need_add_index_df = index_sorted_df[~is_duplicated]
    # print("đây là need_add_index_df : \n",need_add_index_df)
    # thêm các kết quả đã tính toán ở trên vào cuối của mỗi kpi
    added_row = 0
    map_new_index = {}
    for (index, i) in zip(need_add_index_df.index, range(len(merged_sum_df))):
            df_new_record = pd.DataFrame(merged_sum_df.iloc[i, :]).T
            true_position = index+added_row+1
            sorted_df = pd.concat([sorted_df.iloc[:true_position], df_new_record,
                                  sorted_df.iloc[true_position:]]).reset_index(drop=True)
            # thêm các giá trị index mới vào dictionary
            map_new_index[index] = true_position
            added_row += 1
    sorted_df.drop(columns=[None], axis=1, inplace=True)
    sorted_df = sorted_df.apply(lambda x: '' if x.empty else x)
    sorted_df.drop(columns=[type+"_x"], axis=1, inplace=True)
    sorted_df.drop(columns=[type+"_y"], axis=1, inplace=True)
    # thay đổi thành rỗng để có thể groupby phía sau
    sorted_df[tn] = sorted_df[tn].apply(lambda x: '' if x == None else x)
    need_add_index_df.set_index(need_add_index_df.index.map(map_new_index), inplace=True)

    # 1. tạo ra các dataframe chứa tên, level, tên nhóm và dataframe chưa tên các cột  (user_df)
    grouped = sorted_df.groupby([ei, name,email, lv, tn])
    user_df = grouped.size().reset_index(name='Count')
    user_df.insert(0, 'Index', user_df.index+1)
    user_df.drop(columns=['Count'], axis=1, inplace=True)
    # bỏ cột id trong giao diện
    user_df.drop(columns=[ei], axis=1, inplace=True)
    # print("đây là user_df : \n",user_df)

    # Create a new DataFrame to store the column names (title_df)'
    columns_to_drop = [ei, name,email, lv, tn]
    insert_header_name_df = sorted_df.drop(columns=columns_to_drop, axis=1)
    title_df = pd.DataFrame([insert_header_name_df.columns],
                            columns=insert_header_name_df.columns)
    # print("đây là user_df : \n",title_df)

    # 2. tạo ra một list chứa các vị trí cần add các record đó vào bằng cách tìm các vị trí đầu xuất hiện của các record
    # Tạo cột mới để xác định lần xuất hiện đầu tiên của giá trị trùng nhau
    sorted_df['FirstOccurrence'] = ~sorted_df.duplicated(subset=[ei])
    # lấy lần xuất hiện đầu tiên của các giá trị trùng nhau (firstOccurrence_df)
    firstOccurrence_df = sorted_df[sorted_df['FirstOccurrence']].drop(
        columns='FirstOccurrence')
    firstOccurrence_list = firstOccurrence_df.index
    sorted_df = sorted_df.drop(columns='FirstOccurrence', axis=1)


    # Chuyển DataFrame thành một đối tượng Sheet bằng pandas dataframe_to_rows
    user_rows = list(dataframe_to_rows(user_df, index=False, header=False))
    if len(user_rows) != len(firstOccurrence_list):
        raise ValueError("Số lượng dòng cần thêm phải bằng số dự liệu muốn thêm!")
    
    # tạo ra 1 dataframe có nhiều hàng là tên các cột cần add vào sheet
    for i in range(len(user_rows)):
         title_df = pd.concat([title_df, title_df], ignore_index=True)

    # lấy vị trí các cột cần mở rộng khi align bằng openpyxl
    sum_column_names = [tsct, kq, tl, et,rt]
    sum_column_positions = [title_df.columns.get_loc(col_name) for col_name in sum_column_names]

    column_name_rows = list(dataframe_to_rows(title_df, index=False, header=False))
    rows = dataframe_to_rows(insert_header_name_df, index=False, header=False)


# format các trường dữ liệu với workbook của openpyxl
# -----------------------------------------------------------------------------------------------------------
    # tạo một sheet excel mới
    workbook = xlsxwriter.Workbook(options={'nan_inf_to_errors': True})
    xlsxsheet = workbook.add_worksheet()
    # Ghi dữ liệu từ DataFrame vào Workbook
    # Gán tiêu đề cho sheet
    Header_font = Font(name='Arial',size=16, bold=True)
    header_format = workbook.add_format({
    'bold': True,
    'font_name': 'Arial',
    'font_size': 12,
    'align': 'center',
    'valign': 'vcenter',  # Optional: Vertically center the text
    'border': 1,  # Optional: Add a border to the header cells
    'bg_color': 'yellow',  # Optional: Set background color
    })  
    header_text=f"nhóm {level}"
    xlsxsheet.set_header(header_text, {'font': header_format})

    # Merge các ô trong dòng đầu tiên
    xlsxsheet.merge_range(1, 1, 1, title_df.shape[1], 'Merged Header', header_format)

    # Chèn các dòng vào vị trí xác định (từ dòng 2 trở đi)
    cell_format = workbook.add_format({
    'font_name': 'Arial',
    'font_size': 11,
    'bold': True
    })
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, 1):
            xlsxsheet.write(r_idx, c_idx, value,cell_format)

    # thêm các dòng thông tin người dung và title của các trường vào sheet, đồng thời thêm màu, sửa font
    # light_blue_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')

    yellow_cell_format = workbook.add_format()
    yellow_cell_format.set_bg_color('#FFFF99') 

    light_blue_cell_format = workbook.add_format({
    'font_name': 'Arial',
    'font_size': 11,
    'bold': True
    })
    light_blue_cell_format.set_bg_color('#CCE0FF') 
    light_blue_cell_format.set_font_color('#FF0000')

    dark_blue_cell_format = workbook.add_format({
    'font_name': 'Arial',
    'font_size': 11,
    'bold': True
    })
    dark_blue_cell_format.set_bg_color('#6699CC')  
    dark_blue_cell_format.set_font_color('#FFFFFF')

    light_green_cell_format = workbook.add_format()
    light_green_cell_format.set_bg_color('#B0E57C') 

    dark_green_cell_format = workbook.add_format()
    dark_green_cell_format.set_bg_color('#008000')  

    alittler_dark_blue_cell_format = workbook.add_format()
    alittler_dark_blue_cell_format.set_bg_color('#9FC9E7') 
    
    # lấy vị trí các column cần đổi màu xanh lá cây nhạt
    column_light_green_fill = [tsct, kq, tl, et, note]
    column_positions_light_green_fill = [title_df.columns.get_loc(col_name)+1 for col_name in column_light_green_fill]

    # lấy vị trí các column cần đổi màu xanh lá cây đậm
    column_dark_green_fill = [QA]
    column_positions_dark_green_fill = [title_df.columns.get_loc(col_name)+1 for col_name in column_dark_green_fill]

    # # lấy vị trí các column cần đổi màu xanh lá cây đậm
    # column_dark_green_fill = [tsct, 'KR team', 'KR cá nhân', 'Công thức tính', note]
    # column_positions_dark_green_fill = [title_df.columns.get_loc(col_name) for col_name in column_dark_green_fill]

    # lấy vị trí các column cần đổi màu xanh da trời nhạt
    column_light_blue_fill = [type,rt,noteproof]
    column_positions_light_blue_fill = [title_df.columns.get_loc(col_name)+1 for col_name in column_light_blue_fill]

    for row_sum_index in need_add_index_df.index:
        for col_sum_index in sum_column_positions:
            # ở đây row phải +2 là vì: sheet có các index từ số 1 + header là lấy mất 1 dòng đầu
            # ở đây col phải +1 là vì: sheet có các index từ số 1
            xlsxsheet.write_blank(row_sum_index+2, col_sum_index+1, None,yellow_cell_format)

    added_sheet_row = 0
    for insert_index, row_user_data, row_column_name in zip(firstOccurrence_list, user_rows,column_name_rows):       
        user_insert = insert_index+added_sheet_row+2
        title_insert = insert_index+added_sheet_row+3
        # sheet.insert_rows(user_insert, 1)
        # sheet.insert_rows(title_insert, 1)
        # thêm dòng thông tin user
        for col_idx_user,user_value in enumerate(row_user_data, 1):
            xlsxsheet.write(user_insert, col_idx_user, user_value, light_blue_cell_format)
        # thêm dòng title
        for col_idx,column_name_value in enumerate(row_column_name,1):          
            xlsxsheet.write(user_insert, col_idx_user, column_name_value, dark_blue_cell_format)
            
            for green_index in column_positions_light_green_fill:
                if green_index==col_idx:
                    # print("vị trí xanh lá cây:",green_index)
                    xlsxsheet.write_blank(title_insert, green_index, None, light_green_cell_format)

            for dark_green_index in column_positions_dark_green_fill:
                if dark_green_index==col_idx:
                    # print("vị trí xanh lá cây đậm:",dark_green_index)
                    xlsxsheet.write_blank(title_insert, dark_green_index, None, dark_green_cell_format)

            for blue_index in column_positions_light_blue_fill:
                if blue_index==col_idx:
                    # print("vị trí xanh da trời nhạt:",blue_index)
                    xlsxsheet.write_blank(title_insert, blue_index, None, alittler_dark_blue_cell_format)

        if(insert_index!=0):
            #  blank_insert = insert_index+added_sheet_row+2
            #  xlsxsheet.write_blank(title_insert, blue_index, None, alittler_dark_blue_cell_format)
            #  sheet.insert_rows(blank_insert, 1)
             added_sheet_row+=3       
        else:
             added_sheet_row+=2
    # Get the dimensions of the worksheet
    num_rows = xlsxsheet.dim_rowmax + 1
    num_cols = xlsxsheet.dim_colmax + 1
    data_list = []
    # Loop through rows and columns to gather data
    for row_num in range(num_rows):
        row_data = []
        for col_num in range(num_cols):
            cell_value = xlsxsheet.table.get(row_num, col_num)
            print("đây là cell_value:",cell_value)
            row_data.append(cell_value)
        data_list.append(row_data)
    # Biến trạng thái
    merging = False
    merge_start = None
    merge_value = None
    # lấy vị trí cột loại để merge các cell 
    merge_column_position = title_df.columns.get_loc(type)
    # Duyệt qua các dòng trong cột "loại"
    for row_index, row in enumerate(data_list[merge_column_position:], start=1):
        if row[0] =='kpi' or row[0]=='okr':
            if not merging:
                merging = True
                merge_start = row_index
                merge_value = row[0]
        else : 
            if merging:
                merging = False
                xlsxsheet.merge_range(merge_start, 1, row_index - 1, 1, merge_value)
                merge_value=None
    # lấy vị trí column link bằng chứng
    note_proof_column = [noteproof,proofStr,proofURL]
    note_proof_column_position = [title_df.columns.get_loc(col_name) for col_name in note_proof_column]

    hyperlink_format = workbook.add_format({
    'color': 'blue',
    'underline': 1,
    })
    row_index = 0
    for row in enumerate(data_list, start=4):
        text_string = row[note_proof_column_position[1]]  
        url_string = row[note_proof_column_position[2]]   
        row_index += 1 
        if is_valid_string(url_string):      
            if url_string not in proofURL:
                rich_text = CellRichText()
                cell_values = text_string.split(',')  # Split the comma-separated values
                urls = url_string.split(',')
                # print("đây là urls : \n",urls)
                for value, url in zip(cell_values,urls):
                    text_block = workbook.add_rich_string()
                    text_block.append(value)
                    text_block.append('\n', hyperlink_format, url)
                    rich_text.append(text_block)
                    # cell_value = f'=HYPERLINK({url}, {value})'
                xlsxsheet.write(row_index + 3, note_proof_column_position[0]+1, rich_text) 
    # Xóa cột dựa trên vị trí
    xlsxsheet.set_column(note_proof_column_position[1]+1, note_proof_column_position[2]+1, 0)
    # Đặt viền cho các cell
    format_with_border_wrap_center = workbook.add_format({
        'border': 1,          # Border width
        'border_color': 'black',  # Border color
        'text_wrap': True,     # Wrap text
        'align': 'center',     # Center alignment
        'valign': 'vcenter'    # Vertical center alignment
    })
        # Loop through each row in the worksheet
    for row in enumerate(data_list):
        for cell in row:
            xlsxsheet.write_blank(row, cell, None, format_with_border_wrap_center)
    # lấy vị trí các cột cần mở rộng khi align bằng openpyxl
    column_names_30 = [krp, krt, krc, ctt, note,noteproof,QA]
    column_positions_30 = [title_df.columns.get_loc(col_name) for col_name in column_names_30]

    column_names_20 = [ type, et, rt]
    column_positions_20 = [title_df.columns.get_loc(col_name) for col_name in column_names_20]

    # Update the column widths
    for index in column_positions_30:
        xlsxsheet.set_column(index, index, 30)

    # Update the column widths
    for index in column_positions_20:
        xlsxsheet.set_column(index, index, 20)

    # Lưu Workbook
    workbook.close(file_path)
